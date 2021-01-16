try:
    from openpyxl import load_workbook
    
    wb = load_workbook('apriori_data.xlsx')
    ws = wb.active
    
    billing_data = {}
    
    apriori = {}
    
    all_products = {}
    
    for index, row in enumerate(ws.rows):
        if index == 0:
            continue

        index_str = str(index + 1)
        
        bill_no = ws["B" + index_str].value
        product_name = ws["A" + index_str].value
        barcode = str(ws["C" + index_str].value)
        if barcode not in all_products:
            all_products[barcode] = product_name
        if bill_no not in billing_data:
            billing_data[bill_no] = []
            
        billing_data[bill_no].append({'barcode': barcode, 'product_name': product_name})
    
    for bill_no, products in billing_data.items():
        for index, product_data in enumerate(products):
            barcode = product_data['barcode']
            if barcode not in apriori:
                apriori[barcode] = {}
            for product_data_others in billing_data[bill_no]:
                if product_data_others['barcode'] is not product_data['barcode']:
                    if product_data_others['barcode'] not in apriori[product_data['barcode']]:
                         apriori[product_data['barcode']][product_data_others['barcode']] = 1
                    else:
                         apriori[product_data['barcode']][product_data_others['barcode']] += 1
    apriori = dict(apriori)
    save_str = ""
    for barcode, apriority in apriori.items():
        for xbarcode, qty in sorted(apriority.items(), key=lambda item: item[1], reverse=True):
            save_str += all_products[barcode] + " and " + all_products[xbarcode] + " products bought together " + str(qty) + " times.\n"
            print(all_products[barcode] + " and " + all_products[xbarcode] + " products bought together " + str(qty) + " times.")
        save_str += "\n"
        print("")
    
    print("File is being saved...")
    
    file_to_write = open(".\\apriori_processed_data.txt","w+")
    file_to_write.write(save_str)
    file_to_write.close()
    
    print("File is saved successfully!")
    
except Exception as err:
    print(err)
input("")
